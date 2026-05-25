# export.py
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter
from typing import List, Dict
import os

def generate_excel(filepath: str, settings: Dict, employees: List[Dict], 
                   month_data: Dict[int, Dict], year: int, month: int):
    """month_data = {emp_id: {1: {'code':'Ф/Я', 'hours':8}, 2: {...}, ...}}"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Табель_{year}_{month:02d}"
    
    # --- Стили ---
    font_header = Font(name='Arial', size=9, bold=True)
    font_normal = Font(name='Arial', size=9)
    font_small = Font(name='Arial', size=8)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    def style_cell(cell, f=font_normal, b=thin_border, a=left_align):
        cell.font = f; cell.border = b; cell.alignment = a
        
    # --- Шапка ---
    ws.merge_cells('A1:J1')
    ws['A1'].value = settings.get('institution', 'Учреждение')
    style_cell(ws['A1'], font_header, a=center_align)
    
    ws.merge_cells('A2:C2')
    ws['A2'].value = f"ОКУД 504421"
    style_cell(ws['A2'], font_small)
    
    ws.merge_cells('D2:G2')
    ws['D2'].value = f"Вид формы: 0"
    style_cell(ws['D2'], font_small)
    
    ws.merge_cells('H2:J2')
    ws['H2'].value = f"Период: {year}г. {month:02d}мес."
    style_cell(ws['H2'], font_small)
    
    ws.merge_cells('A3:B3')
    ws['A3'].value = "Структурное подразделение:"
    style_cell(ws['A3'], font_small)
    ws['C3'].value = settings.get('department', 'Техотдел')
    style_cell(ws['C3'], font_normal)
    
    # --- Заголовки таблицы ---
    headers = ['№', 'ФИО', 'Таб. №', 'Ставка', 'Должность']
    headers += [str(d) for d in range(1, 32)]
    headers += ['Дни 1-15', 'Часы 1-15', 'Дни 16-31', 'Часы 16-31', 'Всего Дней', 'Всего Часов']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        style_cell(cell, font_header, center_align)
        if col > 5 and col <= 36:
            cell.fill = header_fill
            
    # --- Данные ---
    for r, emp in enumerate(employees, 6):
        eid = emp['id']
        data = month_data.get(eid, {})
        
        ws.cell(row=r, column=1, value=r-5).alignment = center_align
        ws.cell(row=r, column=2, value=emp['full_name']).alignment = left_align
        ws.cell(row=r, column=3, value=emp['tab_number']).alignment = center_align
        ws.cell(row=r, column=4, value=emp['rate']).number_format = '0.00'
        ws.cell(row=r, column=4).alignment = center_align
        pos = f"{emp['position_main']} / {emp['position_part']}" if emp['position_part'] else emp['position_main']
        ws.cell(row=r, column=5, value=pos).alignment = left_align
        
        d15, h15, d16, h16, d_all, h_all = 0, 0, 0, 0, 0, 0
        
        for day in range(1, 32):
            cell = ws.cell(row=r, column=5+day)
            if day in data:
                cell.value = data[day]['code']
                h = data[day]['hours']
                if day <= 15:
                    d15 += 1
                    h15 += h
                else:
                    d16 += 1
                    h16 += h
                d_all += 1
                h_all += h
            cell.alignment = center_cell = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.font = font_normal
            
        # Итоги
        totals = [d15, round(h15, 1), d16, round(h16, 1), d_all, round(h_all, 1)]
        for i, val in enumerate(totals):
            cell = ws.cell(row=r, column=36+i+1)
            cell.value = val
            cell.number_format = '0.00' if 'Час' in headers[36+i] else '0'
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(name='Arial', size=9, bold=True)
            
        # Границы для всей строки
        for col in range(1, 43):
            ws.cell(row=r, column=col).border = thin_border
            
    # --- Подписи ---
    sign_row = r + 3
    signs = [
        ('Ответственный за ведение табеля:', settings.get('responsible', '')),
        ('Начальник тех. отдела:', settings.get('head_dept', '')),
        ('Бухгалтер:', settings.get('accountant', ''))
    ]
    for i, (lbl, val) in enumerate(signs):
        ws.cell(row=sign_row+i, column=1, value=lbl).font = font_normal
        ws.cell(row=sign_row+i, column=3, value=val).font = Font(name='Arial', size=9, italic=True)
        ws.cell(row=sign_row+i, column=1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        
    # --- Настройки листа ---
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 30
    for col in range(6, 37):
        ws.column_dimensions[get_column_letter(col)].width = 4.5
    for col in range(37, 43):
        ws.column_dimensions[get_column_letter(col)].width = 12
        
    ws.freeze_panes = 'F6'
    wb.save(filepath)
    return filepath
